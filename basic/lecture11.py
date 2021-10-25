
# pip install openpyxl
#import tkinter
#import thinter as gg
#from tkinter import *
#k = tkinter.Button(win)

import openpyxl

# wb = openpyxl.load_workbook('text2.xlsx')
# sh = wb.active
# sh.cell(1,3).value='korea'
# sh['A3']='korea'

t= "김인문, 010-2244-3545, 상계동 237번지, abc@naver.com,p"

t= t.split(',')

wb=openpyxl.Workbook()
ws = wb.active

for i ,ele in enumerate(t):
    ws.cell(1, i+1, ele)
wb.save('sameple.xlsx')