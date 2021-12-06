from openpyxl import *

wb = wb2 = load_workbook('lotto/data/lotto46.xlsx');
print(wb.sheetnames)

ws = wb.active

for row in ws.iter_rows(4):
    print( "{0:4d} : {1:2d}|{2:2d}|{3:2d}|{4:2d}|{5:2d}|{6:2d}|{7:2d}".format(row[1].value,row[13].value,row[14].value,row[15].value,row[16].value,row[17].value,row[18].value,row[19].value))