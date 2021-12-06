from openpyxl import *

wb = wb2 = load_workbook('lotto/data/lotto20.xlsx');
print(wb.sheetnames)

ws = wb.active

for row in ws.iter_rows(4):
    print( "{0} : {1}|{2}|{3}|{4}|{5}|{6}|{7}|{8}".format(row[1].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value,row[8].value,row[9].value,row[10].value))