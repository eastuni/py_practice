import openpyxl                  

wb=openpyxl.Workbook()       #엑셀파일 로드
ws=wb.active                             # 워크시트 핸들러

ws['a1']='안녕하세요'                   #셀접근방법 1  
ws.cell(2,1).value = '파이썬'       #셀접근방법 2 
ws.cell(3,1, "개나리")                    #셀접근방법 3
wb.save('sample.xlsx')                 #파일저장
 
