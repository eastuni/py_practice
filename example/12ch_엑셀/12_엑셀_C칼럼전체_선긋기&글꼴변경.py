from openpyxl import *
from openpyxl.styles import  Alignment, Font, Border, Side
wb = Workbook()
ws = wb.active

for i in range(1,10):
    for j in range(1,10):
        ws.cell(i, j, i*j)

c=ws['c']
for col in c:
    col.font=Font(name= "HY헤드라인M", bold=True)
    col.alignment=Alignment(horizontal='center',vertical='center' )  
    th=Side(border_style='thin')
    col.border=Border(top=th, bottom=th, left=th, right=th)

wb.save("BoldDemo2.xlsx")
print("위 엑셀파일을 열어 확인해보세요")
