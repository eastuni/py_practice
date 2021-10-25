from openpyxl import *
from openpyxl.styles import  Alignment, Font, Border, Side
wb = Workbook()
ws = wb.active
ws['B3'] = "Hello"
ws['B3'].font =Font(name= "HY헤드라인M",
                               bold=True,
                               size=20,
                               italic=True,
                               underline='single')
ws['b3'].alignment= Alignment(horizontal='center',
                                                     vertical='center' )     #top, bottom left, right
th=Side(border_style='thin')
db=Side(border_style='double')
ws['b3'].border=Border(top=th, bottom=th, left=db, right=db)

wb.save("BoldDemo.xlsx")

print("볼드및 테두리 완료. 위 엑셀파일 열어서 확인해보세요")
